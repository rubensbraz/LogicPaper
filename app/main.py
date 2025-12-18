import asyncio
import io
import json
import os
import shutil
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

import anyio
import pandas as pd
from fastapi import FastAPI, File, Form, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.core.batch import process_batch_core
from app.core.config import settings, logger
from app.core.engine import DocumentEngine
from app.core.validator import TemplateValidator
from app.utils import extract_zip, sanitize_filename, start_scheduler
from app.integration.router import router as integration_router


# --- App Initialization ---


# This dictionary defines the sections (tags) visible in the Swagger UI
tags_metadata = [
    {
        "name": "Integration (Headless)",
        "description": "Endpoints for system-to-system integration (ERP, CRM) using JSON and API Keys.",
    },
    {
        "name": "Web Dashboard API",
        "description": "Endpoints used by the Frontend UI (index.html) for interactive upload and validation.",
    },
    {
        "name": "Static Pages",
        "description": "Routes that serve the static HTML content (UI).",
    },
]

# Initialize FastAPI with metadata
app = FastAPI(
    title=settings.PROJECT_NAME,
    version=settings.VERSION,
    description="Batch Processing Engine.",
    openapi_tags=tags_metadata,
)

# Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Start Cleanup Scheduler
start_scheduler(settings.TEMP_DIR, settings.CLEANUP_INTERVAL_SECONDS)


# --- Register Routers ---


app.include_router(
    integration_router,
    prefix=f"{settings.API_PREFIX}/integration",
    tags=["Integration (Headless)"],
)


# --- Real-time Logging (SSE) ---


# Global Event Queue for SSE. In prod, change this for Redis Pub/Sub
log_queues = {}


async def log_generator(session_id: str):
    """
    Async generator for Server-Sent Events (SSE).
    """
    queue = asyncio.Queue()
    log_queues[session_id] = queue

    try:
        while True:
            try:
                # Add a timeout to prevent stale queues from hanging forever
                data = await asyncio.wait_for(
                    queue.get(), timeout=settings.LIBREOFFICE_TIMEOUT
                )
                yield f"data: {data}\n\n"

                if "PROCESS_COMPLETE" in data or "PROCESS_ERROR" in data:
                    break
            except asyncio.TimeoutError:
                yield f"data: HEARTBEAT_TIMEOUT\n\n"
                break
    except Exception as e:
        logger.error(f"SSE Stream error for session {session_id}: {e}")
    finally:
        # Crucial for preventing Memory Leaks
        log_queues.pop(session_id, None)
        logger.info(f"SSE Queue cleared for session: {session_id}")


def send_log(session_id: str, message: str) -> None:
    """Push a log message to the specific session queue."""
    if session_id in log_queues:
        log_queues[session_id].put_nowait(message)


# --- Helper: Data Loader ---


async def load_dataframe(
    file_excel: Optional[UploadFile] = None, file_json: Optional[UploadFile] = None
) -> pd.DataFrame:
    """
    Loads data from either Excel or JSON into a Pandas DataFrame.
    """
    if not file_excel and not file_json:
        raise ValueError("No data source provided. Please upload Excel or JSON.")

    # 1. Handle Excel
    if file_excel:
        try:
            contents = await file_excel.read()
            await file_excel.seek(0)
            return pd.read_excel(io.BytesIO(contents), header=0)
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {str(e)}")

    # 2. Handle JSON
    if file_json:
        try:
            contents = await file_json.read()
            await file_json.seek(0)
            data = json.loads(contents)

            # Case A: Single Dictionary -> Wrap in list
            if isinstance(data, dict):
                data = [data]

            # Case B: List of Dictionaries (Standard)
            elif isinstance(data, list):
                # Validation: Ensure all items are dicts
                if not all(isinstance(i, dict) for i in data):
                    raise ValueError(
                        "JSON list must contain objects (key-value pairs)."
                    )
            else:
                raise ValueError("JSON must be an Object or a List of Objects.")

            # Normalize semi-structured JSON. This flattens simple nested keys if necessary
            return pd.json_normalize(data)

        except json.JSONDecodeError:
            raise ValueError("Invalid JSON file format.")
        except Exception as e:
            raise ValueError(f"Error parsing JSON data: {str(e)}")


# --- Reporting Engine ---


def generate_styled_report(
    path: str,
    report_data: List[Dict[str, Any]],
    metadata: Dict[str, Any],
    input_manifest: Dict[str, Any],
) -> None:
    """
    Generates a high-fidelity, styled Excel report with two sheets.

    Args:
        path (str): Output path for the .xlsx file.
        report_data (List[Dict]): The list of row results (one per file).
        metadata (Dict): Statistics like start_time, duration, file_counts.
        input_manifest (Dict): Dictionary listing input filenames.
    """
    wb = Workbook()

    # --- Styles ---
    navy_blue = "1F4E78"
    white = "FFFFFF"
    light_gray = "F2F2F2"  # Color for alternating rows

    header_font = Font(name="Calibri", size=12, bold=True, color=white)
    title_font = Font(name="Calibri", size=18, bold=True, color=navy_blue)
    subtitle_font = Font(name="Calibri", size=14, bold=True, color=navy_blue)
    label_font = Font(name="Calibri", size=11, bold=True, color="555555")

    header_fill = PatternFill(
        start_color=navy_blue, end_color=navy_blue, fill_type="solid"
    )
    even_fill = PatternFill(
        start_color=light_gray, end_color=light_gray, fill_type="solid"
    )
    success_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    success_font = Font(color="006100")
    error_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    error_font = Font(color="9C0006")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    # Standard Alignment (Wrap + Vertical Center)
    standard_align = Alignment(vertical="center", wrap_text=True)
    # Right Alignment for Column C
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    # Top Alignment for Lists
    top_align = Alignment(vertical="top", wrap_text=True)

    # ==========================
    # SHEET 1: DASHBOARD
    # ==========================
    ws_dash = wb.active
    ws_dash.title = "Executive Summary"
    ws_dash.sheet_view.showGridLines = False

    ws_dash["B2"] = "LogicPaper Execution Report"
    ws_dash["B2"].font = title_font

    # Metrics
    metrics = [
        ("Session ID", metadata["session_id"]),
        ("Date", metadata["start_time"].strftime("%Y-%m-%d")),
        ("Duration", str(metadata["duration"]).split(".")[0]),
        ("Total Rows Processed", metadata["total_rows"]),
        ("Total Files Generated", metadata["total_files"]),
        ("Success Rate", f"{metadata['success_rate']:.1f}%"),
    ]

    row_idx = 4
    for i, (label, value) in enumerate(metrics):
        cell_label = ws_dash.cell(row=row_idx, column=2, value=label)
        cell_value = ws_dash.cell(row=row_idx, column=3, value=value)

        cell_label.font = label_font
        cell_label.border = thin_border
        cell_label.alignment = standard_align

        cell_value.border = thin_border
        cell_value.alignment = right_align

        # Apply alternating color to even rows (relative to the list)
        if i % 2 == 1:
            cell_label.fill = even_fill
            cell_value.fill = even_fill

        row_idx += 1

    # Input Manifest
    row_idx += 2
    ws_dash.cell(row=row_idx, column=2, value="Input Files Manifest").font = (
        subtitle_font
    )
    row_idx += 1

    inputs = [
        ("Data Source", input_manifest.get("excel", "N/A")),
        ("Assets Archive", input_manifest.get("assets", "None")),
        ("Templates Used", "\n".join(input_manifest.get("templates", []))),
    ]

    for i, (label, value) in enumerate(inputs):
        cell_label = ws_dash.cell(row=row_idx, column=2, value=label)
        cell_value = ws_dash.cell(row=row_idx, column=3, value=value)

        cell_label.font = label_font
        cell_label.border = thin_border
        cell_label.alignment = top_align

        cell_value.border = thin_border
        cell_value.alignment = right_align

        # Apply alternating color to even rows (relative to the list)
        if i % 2 == 1:
            cell_label.fill = even_fill
            cell_value.fill = even_fill

        row_idx += 1

    ws_dash.column_dimensions["B"].width = 25
    ws_dash.column_dimensions["C"].width = 35

    # ==========================
    # SHEET 2: DETAILED LOG
    # ==========================
    ws_log = wb.create_sheet("Detailed Logs")

    if report_data:
        df = pd.DataFrame(report_data)
        # Explicit column order
        cols = ["Row", "Identifier", "Output File", "Status", "Error Details"]
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        df = df[cols]

        # Header
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws_log.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Data
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 2
        ):
            for c_idx, value in enumerate(row, 1):
                cell = ws_log.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = standard_align

                # Status Coloring (Column 4)
                if c_idx == 4:
                    if str(value).lower() == "success":
                        cell.fill = success_fill
                        cell.font = success_font
                    else:
                        cell.fill = error_fill
                        cell.font = error_font

                # Error Details Coloring (Column 5)
                if c_idx == 5 and value:
                    cell.font = error_font

        # Column Widths
        ws_log.column_dimensions["A"].width = 10  # Row
        ws_log.column_dimensions["B"].width = 25  # Identifier
        ws_log.column_dimensions["C"].width = 50  # Output File
        ws_log.column_dimensions["D"].width = 15  # Status
        ws_log.column_dimensions["E"].width = 60  # Error Details

        # Table
        tab = Table(displayName="LogTable", ref=f"A1:E{len(df)+1}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws_log.add_table(tab)

    wb.save(path)


# --- Web Dashboard Endpoints ---


@app.get("/stream-logs/{session_id}", tags=["Web Dashboard API"])
async def stream_logs(session_id: str):
    return StreamingResponse(log_generator(session_id), media_type="text/event-stream")


@app.post("/api/preview", tags=["Web Dashboard API"])
async def preview_data(
    file_excel: UploadFile = File(None), file_json: UploadFile = File(None)
):
    """
    Parses the Excel OR JSON file and returns headers and first 5 rows.
    """
    try:
        # Load Data using Helper
        df = await load_dataframe(file_excel, file_json)

        # Row 1: Headers
        headers = df.columns.tolist()

        # Data: Head 5
        data_rows = df.head(5)

        # Format Preview Data (Raw)
        preview_data = []
        for _, row in data_rows.iterrows():
            # Convert to dict, handle NaNs
            row_dict = row.where(pd.notnull(row), None).to_dict()
            # Convert non-serializable types for JSON
            for k, v in row_dict.items():
                if isinstance(v, (datetime, pd.Timestamp)):
                    row_dict[k] = str(v)
            preview_data.append(row_dict)

        return JSONResponse(
            {"status": "success", "headers": headers, "preview": preview_data}
        )
    except Exception as e:
        logger.error(f"Preview failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)


@app.post("/api/validate", tags=["Web Dashboard API"])
async def validate_compatibility(
    file_excel: UploadFile = File(None),
    file_json: UploadFile = File(None),
    files_templates: List[UploadFile] = File(...),
):
    """
    Validates that template tags exist in Excel/JSON headers.
    """
    session_id = f"val_{uuid.uuid4().hex[:8]}"
    session_path = os.path.join(settings.TEMP_DIR, session_id)
    os.makedirs(session_path, exist_ok=True)

    try:
        # Load Data Headers (No need to save file to disk just for headers)
        df = await load_dataframe(file_excel, file_json)
        headers = [str(h) for h in df.columns.tolist()]

        # Save Templates
        templates_map = {}
        for tmpl in files_templates:
            t_path = os.path.join(session_path, tmpl.filename)
            async with await anyio.open_file(t_path, "wb") as f:
                await f.write(await tmpl.read())
            templates_map[tmpl.filename] = t_path

        # Validate
        validator = TemplateValidator()
        result = validator.compare(headers, templates_map)

        return JSONResponse({"status": "success", "report": result})

    except Exception as e:
        logger.error(f"Validation Error: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        if os.path.exists(session_path):
            shutil.rmtree(session_path)


@app.post("/api/sample", tags=["Web Dashboard API"])
async def generate_sample(
    session_id: str = Form(...),
    filename_col: str = Form(...),
    output_pdf: bool = Form(False),
    file_excel: UploadFile = File(None),
    file_json: UploadFile = File(None),
    files_templates: List[UploadFile] = File(...),
    file_assets: UploadFile = File(None),
):
    """
    Dry Run Endpoint: Processes ONLY the first data row from the Excel/JSON file.
    Returns a ZIP file containing the generated documents for that specific row
    for immediate user verification.
    """
    start_time = datetime.now()

    # 1. Setup Temporary Session
    sample_session_id = f"{session_id}_sample"
    session_path = os.path.join(settings.TEMP_DIR, sample_session_id)

    dir_inputs = os.path.join(session_path, "inputs")
    dir_output = os.path.join(session_path, "output")
    dir_assets_internal = os.path.join(session_path, ".temp_assets")

    for p in [dir_inputs, dir_output, dir_assets_internal]:
        os.makedirs(p, exist_ok=True)

    try:
        # 2. Load and Save Input Data
        df = await load_dataframe(file_excel, file_json)
        input_filename = file_excel.filename if file_excel else file_json.filename

        source_upload = file_excel if file_excel else file_json
        if source_upload:
            # Ensures the cursor is at the beginning (because the load_dataframe has already read the file)
            await source_upload.seek(0)
            source_path = os.path.join(dir_inputs, source_upload.filename)

            # Saves the original file to the inputs folder
            async with await anyio.open_file(source_path, "wb") as f:
                await f.write(await source_upload.read())

            # Reset the cursor again
            await source_upload.seek(0)

        # Save Templates
        saved_template_paths = []
        template_names = []
        for tmpl in files_templates:
            t_path = os.path.join(dir_inputs, tmpl.filename)
            async with await anyio.open_file(t_path, "wb") as f:
                await f.write(await tmpl.read())
            saved_template_paths.append(t_path)
            template_names.append(tmpl.filename)

        # Handle Assets
        assets_filename = "None"
        if file_assets:
            assets_filename = file_assets.filename
            zip_input_path = os.path.join(dir_inputs, file_assets.filename)
            async with await anyio.open_file(zip_input_path, "wb") as f:
                await f.write(await file_assets.read())
            extract_zip(zip_input_path, dir_assets_internal)

        # Manifest for Report
        input_manifest = {
            "excel": input_filename,
            "templates": template_names,
            "assets": assets_filename,
        }

        # 3. Initialize Engine & Parse Data
        engine = DocumentEngine(session_path)

        if df.empty:
            raise ValueError("Data source is empty.")

        # TARGET: Only the first data row (Index 0, which corresponds to Excel Row 2)
        target_row = df.iloc[0]

        # 4. Process Single Row
        report = []
        files_generated_count = 0
        row_success = False

        row_identifier = "SAMPLE"

        # Parse Context (Raw Data)
        context = target_row.to_dict()

        # Sanitize Context (Handle NaN and NaT)
        cleaned_context = {}
        try:
            for k, v in context.items():
                if pd.isna(v):
                    cleaned_context[k] = None
                else:
                    cleaned_context[k] = v

            # Determine Identifier
            if filename_col in cleaned_context and cleaned_context[filename_col]:
                row_identifier = sanitize_filename(str(cleaned_context[filename_col]))

        except Exception as e:
            report.append(
                {
                    "Row": 1,
                    "Identifier": "SAMPLE",
                    "Output File": "N/A",
                    "Status": "Error",
                    "Error Details": f"Data Parsing Error: {str(e)}",
                }
            )

        # Generate Files (if no parsing error)
        if not report:
            for tmpl_path in saved_template_paths:
                tmpl_filename = os.path.basename(tmpl_path)
                tmpl_name_base, tmpl_ext = os.path.splitext(tmpl_filename)

                final_filename = (
                    f"{tmpl_name_base} - {row_identifier} (SAMPLE){tmpl_ext}"
                )
                doc_output_path = os.path.join(dir_output, final_filename)

                try:
                    if tmpl_ext.lower() == ".docx":
                        await engine.process_docx(
                            tmpl_path,
                            doc_output_path,
                            cleaned_context,
                            dir_assets_internal,
                        )
                    elif tmpl_ext.lower() in [".md", ".txt"]:
                        await engine.process_text(
                            tmpl_path, doc_output_path, cleaned_context
                        )
                    elif tmpl_ext.lower() == ".pptx":
                        await engine.process_pptx(
                            tmpl_path, doc_output_path, cleaned_context
                        )

                    report.append(
                        {
                            "Row": 1,
                            "Identifier": row_identifier,
                            "Output File": final_filename,
                            "Status": "Success",
                            "Error Details": "",
                        }
                    )
                    files_generated_count += 1
                    row_success = True

                except Exception as e:
                    report.append(
                        {
                            "Row": 1,
                            "Identifier": row_identifier,
                            "Output File": final_filename,
                            "Status": "Error",
                            "Error Details": str(e),
                        }
                    )

                # PDF Conversion
                if output_pdf and os.path.exists(doc_output_path):
                    pdf_filename = f"{tmpl_name_base} - {row_identifier} (SAMPLE).pdf"
                    try:
                        await engine.convert_to_pdf(doc_output_path, dir_output)
                        report.append(
                            {
                                "Row": 1,
                                "Identifier": row_identifier,
                                "Output File": pdf_filename,
                                "Status": "Success",
                                "Error Details": "",
                            }
                        )
                        files_generated_count += 1
                    except Exception as e:
                        report.append(
                            {
                                "Row": 1,
                                "Identifier": row_identifier,
                                "Output File": pdf_filename,
                                "Status": "Error",
                                "Error Details": f"PDF Error: {str(e)}",
                            }
                        )

        # 5. Generate Report & Zip
        end_time = datetime.now()
        duration = end_time - start_time

        metadata = {
            "session_id": "SAMPLE_RUN",
            "start_time": start_time,
            "duration": duration,
            "total_rows": 1,
            "total_files": files_generated_count,
            "success_rate": 100 if row_success else 0,
        }

        # Generate Styled Report inside Output Folder
        report_path = os.path.join(dir_output, "sample_report.xlsx")
        generate_styled_report(report_path, report, metadata, input_manifest)

        # Clean assets
        if os.path.exists(dir_assets_internal):
            shutil.rmtree(dir_assets_internal)

        # Zip Output
        zip_base_name = os.path.join(settings.TEMP_DIR, f"{session_id}_sample_result")
        shutil.make_archive(base_name=zip_base_name, format="zip", root_dir=dir_output)

        zip_file_path = f"{zip_base_name}.zip"
        timestamp = end_time.strftime("%Y-%m-%d_%H-%M")
        download_filename = f"LogicPaper_Sample_{row_identifier}_{timestamp}.zip"

        return FileResponse(
            path=zip_file_path, filename=download_filename, media_type="application/zip"
        )

    except Exception as e:
        logger.error(f"Sample Generation Error: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)


@app.post("/api/process", tags=["Web Dashboard API"])
async def process_batch(
    session_id: str = Form(...),
    filename_col: str = Form(...),
    output_pdf: bool = Form(False),
    group_by_folders: bool = Form(True),
    file_excel: UploadFile = File(None),
    file_json: UploadFile = File(None),
    files_templates: List[UploadFile] = File(...),
    file_assets: UploadFile = File(None),
):
    """
    Main batch processing endpoint.
    """
    start_time = datetime.now()
    session_path = os.path.join(settings.TEMP_DIR, session_id)

    dir_inputs = os.path.join(session_path, "1 Input documents")
    dir_outputs = os.path.join(session_path, "2 Generated documents")
    dir_assets_internal = os.path.join(session_path, ".temp_assets")

    for p in [dir_inputs, dir_outputs, dir_assets_internal]:
        os.makedirs(p, exist_ok=True)

    try:
        send_log(session_id, "🟢 Initializing session structure...")

        # 1. Load Data
        df = await load_dataframe(file_excel, file_json)
        input_filename = file_excel.filename if file_excel else file_json.filename

        # Save Source File
        source_upload = file_excel if file_excel else file_json
        if source_upload:
            await source_upload.seek(0)
            source_path = os.path.join(dir_inputs, source_upload.filename)
            async with await anyio.open_file(source_path, "wb") as f:
                await f.write(await source_upload.read())
            await source_upload.seek(0)

        # 2. Save Templates
        saved_template_paths = []
        template_names = []
        for tmpl in files_templates:
            t_path = os.path.join(dir_inputs, tmpl.filename)
            async with await anyio.open_file(t_path, "wb") as f:
                await f.write(await tmpl.read())
            saved_template_paths.append(t_path)
            template_names.append(tmpl.filename)
            send_log(session_id, f"   ↳ Loaded template: {tmpl.filename}")

        # 3. Handle Assets
        assets_filename = "None"
        if file_assets:
            assets_filename = file_assets.filename
            zip_input_path = os.path.join(dir_inputs, file_assets.filename)
            async with await anyio.open_file(zip_input_path, "wb") as f:
                await f.write(await file_assets.read())
            extract_zip(zip_input_path, dir_assets_internal)
            send_log(session_id, "📦 Assets library prepared.")

        total_rows = len(df)
        send_log(session_id, f"🚀 Starting processing for {total_rows} rows.")

        # 4. Execute Core batch logic

        # Define callback wrapper for SSE
        def sse_callback(msg: str):
            send_log(session_id, msg)

        batch_result = await process_batch_core(
            session_id=session_id,
            df=df,
            template_paths=saved_template_paths,
            session_path=session_path,
            dir_outputs=dir_outputs,
            dir_assets=dir_assets_internal,
            to_pdf=output_pdf,
            filename_col=filename_col,
            group_folders=group_by_folders,
            log_callback=sse_callback,
        )

        # 5. Cleanup & Report
        if os.path.exists(dir_assets_internal):
            shutil.rmtree(dir_assets_internal)

        end_time = datetime.now()
        duration = end_time - start_time

        input_manifest = {
            "excel": input_filename,
            "templates": template_names,
            "assets": assets_filename,
        }

        metadata = {
            "session_id": session_id,
            "start_time": start_time,
            "duration": duration,
            "total_rows": batch_result["total_rows"],
            "total_files": batch_result["total_files"],
            "success_rate": (
                (batch_result["success_rows"] / batch_result["total_rows"] * 100)
                if batch_result["total_rows"] > 0
                else 0
            ),
        }

        report_path = os.path.join(session_path, "job_report.xlsx")
        generate_styled_report(
            report_path, batch_result["report"], metadata, input_manifest
        )

        zip_base_name = os.path.join(settings.TEMP_DIR, f"{session_id}_result")
        shutil.make_archive(
            base_name=zip_base_name, format="zip", root_dir=session_path
        )

        send_log(session_id, "PROCESS_COMPLETE")
        return JSONResponse(
            {"status": "success", "download_url": f"/api/download/{session_id}"}
        )

    except Exception as e:
        logger.error(f"Critical Batch Error: {e}")
        send_log(session_id, f"PROCESS_ERROR: {str(e)}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)


@app.get("/api/download/{session_id}", tags=["Web Dashboard API"])
async def download_result(session_id: str) -> Any:
    """
    Downloads the final ZIP file with a timestamped filename.

    Format: LogicPaper_YYYY-MM-DD_HH-MM.zip
    """
    try:
        file_path = os.path.join(settings.TEMP_DIR, f"{session_id}_result.zip")

        if os.path.exists(file_path):
            # Get current time
            now = datetime.now()
            # Format: YYYY-MM-DD_HH-MM
            timestamp = now.strftime("%Y-%m-%d_%H-%M")
            filename = f"LogicPaper_{timestamp}.zip"

            return FileResponse(
                path=file_path, filename=filename, media_type="application/zip"
            )

        return JSONResponse(
            {"status": "error", "message": "File not found"}, status_code=404
        )

    except Exception as e:
        logger.error(f"Download Error: {e}")
        return JSONResponse(
            {"status": "error", "message": "Internal Server Error during download"},
            status_code=500,
        )


# --- Static Pages ---


@app.get("/", tags=["Static Pages"])
async def read_root():
    """Serves the main application page."""
    return FileResponse(os.path.join(settings.STATIC_DIR, "index.html"))


@app.get("/help", tags=["Static Pages"])
async def read_help():
    """Serves the documentation page."""
    return FileResponse(os.path.join(settings.STATIC_DIR, "help.html"))


# --- STATIC FILES CONFIGURATION (SPA/Static Site Mode) ---


# This mounts the 'static' folder to the root URL ("/")
# It allows relative paths (e.g., "css/style.css") to work locally AND on GitHub Pages
# 'html=True' automatically serves 'index.html' when accessing root
app.mount("/", StaticFiles(directory=settings.STATIC_DIR, html=True), name="site")
