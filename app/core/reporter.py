import io
import logging
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configure Logging
logger = logging.getLogger(__name__)


def generate_styled_report(
    report_data: List[Dict[str, Any]],
    metadata: Dict[str, Any],
    input_manifest: Dict[str, Any],
) -> bytes:
    """Generates a high-fidelity, styled Excel report with two sheets.

    Args:
        report_data (List[Dict[str, Any]]): The list of row results (one per file).
        metadata (Dict[str, Any]): Statistics like start_time, duration, file_counts.
        input_manifest (Dict[str, Any]): Dictionary listing input filenames.
    """
    wb = Workbook()

    # --- Styles ---
    navy_blue = "1F4E78"
    white = "FFFFFF"
    light_gray = "F2F2F2"  # Color for alternating rows

    header_font = Font(name="Calibri", size=12, bold=True, color=white)
    title_font = Font(name="Calibri", size=18, bold=True, color=navy_blue)
    subtitle_font = Font(name="Calibri", size=14, bold=True, color=navy_blue)
    label_font = Font(name="Calibri", size=11, bold=True, color="555555")

    header_fill = PatternFill(
        start_color=navy_blue, end_color=navy_blue, fill_type="solid"
    )
    even_fill = PatternFill(
        start_color=light_gray, end_color=light_gray, fill_type="solid"
    )
    success_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    success_font = Font(color="006100")
    error_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    error_font = Font(color="9C0006")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    # Standard Alignment (Wrap + Vertical Center)
    standard_align = Alignment(vertical="center", wrap_text=True)
    # Right Alignment for Column C
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    # Top Alignment for Lists
    top_align = Alignment(vertical="top", wrap_text=True)

    # --- Sheet 1: Executive Summary ---
    ws_dash = wb.active
    ws_dash.title = "Executive Summary"
    ws_dash.sheet_view.showGridLines = False

    ws_dash["B2"] = "LogicPaper Execution Report"
    ws_dash["B2"].font = title_font

    # Metrics
    metrics = [
        ("Session ID", metadata["session_id"]),
        ("Date", metadata["start_time"].strftime("%Y-%m-%d")),
        ("Duration", str(metadata["duration"]).split(".")[0]),
        ("Total Rows Processed", metadata["total_rows"]),
        ("Total Files Generated", metadata["total_files"]),
        ("Success Rate", f"{metadata['success_rate']:.1f}%"),
    ]

    row_idx = 4
    for i, (label, value) in enumerate(metrics):
        cell_label = ws_dash.cell(row=row_idx, column=2, value=label)
        cell_value = ws_dash.cell(row=row_idx, column=3, value=value)

        cell_label.font = label_font
        cell_label.border = thin_border
        cell_label.alignment = standard_align

        cell_value.border = thin_border
        cell_value.alignment = right_align

        # Apply alternating color to even rows (relative to the list)
        if i % 2 == 1:
            cell_label.fill = even_fill
            cell_value.fill = even_fill

        row_idx += 1

    # Input Manifest
    row_idx += 2
    ws_dash.cell(
        row=row_idx, column=2, value="Input Files Manifest"
    ).font = subtitle_font
    row_idx += 1

    inputs = [
        ("Data Source", input_manifest.get("excel", "N/A")),
        ("Assets Archive", input_manifest.get("assets", "None")),
        ("Templates Used", "\n".join(input_manifest.get("templates", []))),
    ]

    for i, (label, value) in enumerate(inputs):
        cell_label = ws_dash.cell(row=row_idx, column=2, value=label)
        cell_value = ws_dash.cell(row=row_idx, column=3, value=value)

        cell_label.font = label_font
        cell_label.border = thin_border
        cell_label.alignment = top_align

        cell_value.border = thin_border
        cell_value.alignment = right_align

        # Apply alternating color to even rows (relative to the list)
        if i % 2 == 1:
            cell_label.fill = even_fill
            cell_value.fill = even_fill

        row_idx += 1

    ws_dash.column_dimensions["B"].width = 25
    ws_dash.column_dimensions["C"].width = 35

    # --- Sheet 2: Detailed Logs ---
    ws_log = wb.create_sheet("Detailed Logs")

    if report_data:
        df = pd.DataFrame(report_data)
        # Explicit column order
        cols = ["Row", "Identifier", "Output File", "Status", "Error Details"]
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        df = df[cols]

        # Header
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws_log.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Data
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 2
        ):
            for c_idx, value in enumerate(row, 1):
                cell = ws_log.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = standard_align

                # Status Coloring (Column 4)
                if c_idx == 4:
                    if str(value).lower() == "success":
                        cell.fill = success_fill
                        cell.font = success_font
                    else:
                        cell.fill = error_fill
                        cell.font = error_font

                # Error Details Coloring (Column 5)
                if c_idx == 5 and value:
                    cell.font = error_font

        # Column Widths
        ws_log.column_dimensions["A"].width = 10  # Row
        ws_log.column_dimensions["B"].width = 25  # Identifier
        ws_log.column_dimensions["C"].width = 50  # Output File
        ws_log.column_dimensions["D"].width = 15  # Status
        ws_log.column_dimensions["E"].width = 60  # Error Details

        # Table
        tab = Table(displayName="LogTable", ref=f"A1:E{len(df) + 1}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws_log.add_table(tab)

    # Save to memory buffer
    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    return out_buffer.getvalue()
